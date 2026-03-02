# Azure Functions用のメインファイル
import azure.functions as func
from openpyxl import load_workbook
from datetime import datetime, date
import pandas as pd
import tempfile
import os
import io
import logging
import json
from werkzeug.formparser import parse_form_data
from werkzeug.datastructures import Headers

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Azure Functions アプリケーションの初期化
app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

# CORSヘッダーを追加するヘルパー関数
def add_cors_headers(response: func.HttpResponse) -> func.HttpResponse:
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route(route="", methods=["GET"])
def home(req: func.HttpRequest) -> func.HttpResponse:
    """
    ヘルスチェック用のルートエンドポイント
    """
    logging.info("Health check endpoint accessed")
    response = func.HttpResponse(
        body='{"status": "OK", "message": "Timesheet API is running"}',
        mimetype="application/json",
        status_code=200
    )
    return add_cors_headers(response)

@app.route(route="upload", methods=["POST", "OPTIONS"])
def generate_timesheet(req: func.HttpRequest) -> func.HttpResponse:
    """
    メインのタイムシート生成機能
    CSV2個とExcel1個のファイルをアップロードして、勤怠データを処理し
    Excelテンプレートに反映して新しいファイルを生成する
    """
    
    # OPTIONSリクエスト（CORSプリフライト）への対応
    if req.method == "OPTIONS":
        response = func.HttpResponse(status_code=200)
        return add_cors_headers(response)
    
    try:
        logging.info("=== Starting generate_timesheet ===")
        logging.info(f"Request method: {req.method}")
        content_type = req.headers.get('Content-Type', '')
        logging.info(f"Content-Type: {content_type}")
        
        # werkzeugを使ってフォームデータをパース
        body = req.get_body()
        logging.info(f"Body length: {len(body)}")
        
        environ = {
            'REQUEST_METHOD': req.method,
            'CONTENT_TYPE': content_type,
            'CONTENT_LENGTH': str(len(body)),
            'wsgi.input': io.BytesIO(body),
            'wsgi.errors': io.BytesIO(),
            'wsgi.url_scheme': 'https',
            'SERVER_NAME': 'timesheet-api-func-free.azurewebsites.net',
            'SERVER_PORT': '443',
            'PATH_INFO': '/api/upload'
        }
        
        try:
            stream, form, files = parse_form_data(environ)
            logging.info(f"Form parsed successfully")
            logging.info(f"Form keys: {list(form.keys())}")
            logging.info(f"Files keys: {list(files.keys())}")
        except Exception as parse_error:
            logging.error(f"Error parsing form data: {parse_error}")
            logging.exception("Parse traceback:")
            raise
        
        # ファイルの取得
        uploaded_files = files.getlist("files")
        logging.info(f"Files received: {len(uploaded_files)}")
        
        # フォームデータの取得
        name = form.get("name")
        eid = form.get("eid")
        organization = form.get("organization")
        logging.info(f"Form data - name: {name}, org: {organization}")
        year = int(form.get("year"))
        month = int(form.get("month"))
        logging.info(f"Date - year: {year}, month: {month}")
        task = form.get("task")
        time_mode = form.get("time_mode", "none")  # "none", "ratio", "fixed"
        ratio_day_fraction = None
        fixed_hours_fraction = None
        additional_breaks = []

        # 時間モードのバリデーション
        if time_mode == "ratio":
            ratio_percent_raw = form.get("ratio_percent")
            try:
                ratio_percent = float(ratio_percent_raw)
            except (TypeError, ValueError):
                response = func.HttpResponse("就業時間割合は0〜100の数値で指定してください", status_code=400)
                return add_cors_headers(response)

            if ratio_percent < 0 or ratio_percent > 100:
                response = func.HttpResponse("就業時間割合は0〜100の範囲で指定してください", status_code=400)
                return add_cors_headers(response)

            ratio_hours = 8.0 * ratio_percent / 100.0
            ratio_day_fraction = ratio_hours / 24.0

        elif time_mode == "fixed":
            fixed_hours_raw = form.get("fixed_hours")
            try:
                fixed_hours = float(fixed_hours_raw)
            except (TypeError, ValueError):
                response = func.HttpResponse("固定時間は0以上の数値で指定してください", status_code=400)
                return add_cors_headers(response)

            if fixed_hours < 0:
                response = func.HttpResponse("固定時間は0以上で指定してください", status_code=400)
                return add_cors_headers(response)

            fixed_hours_fraction = fixed_hours / 24.0

        # 追加の休憩時間を取得
        additional_breaks_json = form.get("additional_breaks")
        if additional_breaks_json:
            try:
                additional_breaks = json.loads(additional_breaks_json)
            except json.JSONDecodeError:
                additional_breaks = []

        # アップロードファイルの分類と検証
        logging.info("Checking CSV files...")
        csv_files = [f for f in uploaded_files if f.filename.lower().endswith(".csv")]
        logging.info(f"CSV files found: {len(csv_files)}")
        if len(csv_files) != 2:
            logging.warning(f"Invalid number of CSV files: {len(csv_files)}")
            response = func.HttpResponse("CSV2個をアップロードしてください", status_code=400)
            return add_cors_headers(response)

        # テンプレートファイルのパスを設定
        logging.info("Checking template file...")
        template_path = os.path.join(BASE_DIR, "templates", "Excel_templates", "タイムシート(yyyy_mm).xlsx")
        logging.info(f"Template path: {template_path}")
        logging.info(f"BASE_DIR: {BASE_DIR}")
        logging.info(f"Template exists: {os.path.exists(template_path)}")
        if not os.path.exists(template_path):
            logging.error("Template file not found!")
            response = func.HttpResponse("テンプレートファイルが見つかりません", status_code=500)
            return add_cors_headers(response)

        # CSV読み込み＆データ結合処理
        logging.info("Reading CSV files...")
        df_all = pd.concat([pd.read_csv(f.stream) for f in csv_files], ignore_index=True)
        logging.info(f"Total rows: {len(df_all)}")
        logging.info(f"Columns: {df_all.columns.tolist()}")
        
        # 日時データの型変換（エラーの場合はNaTに変換）
        df_all["Work start"] = pd.to_datetime(df_all["Work start"], errors="coerce")
        df_all["Work end"] = pd.to_datetime(df_all["Work end"], errors="coerce")
        df_all["Break start"] = pd.to_datetime(df_all["Break start"], errors="coerce")
        df_all["Break end"] = pd.to_datetime(df_all["Break end"], errors="coerce")
        
        # 日付列を作成し、開始時刻順にソート
        df_all["Date"] = df_all["Work start"].dt.date
        df_all.sort_values("Work start", inplace=True)

        # Excelテンプレートファイルの読み込み
        wb = load_workbook(filename=template_path)
        ws = wb.worksheets[0]
        ws.title = f"{month}月"

        # 月の日数を設定
        days_in_month = pd.Timestamp(year=year, month=month, day=1).days_in_month

        # 基本情報の設定
        d9_date = date(year, month, 1)
        g9_date = date(year, month, days_in_month)
        ws["D6"] = organization
        ws["D8"] = name
        ws["D9"] = d9_date
        ws["G9"] = g9_date

        # 祝日データの読み込み
        holiday_dict = {}
        try:
            with open(os.path.join(BASE_DIR, "holidays.csv"), encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line: continue
                    parts = line.split(",")
                    if len(parts) >= 2:
                        holiday_dict[parts[0]] = parts[1]
        except:
            pass

        # 各日のデータ処理ループ
        for day in range(1, days_in_month + 1):
            current_date = date(year, month, day)
            row = 12 + day
            date_str = current_date.strftime("%Y-%m-%d")
            day_data = df_all[df_all["Date"] == current_date]
            holiday_name = holiday_dict.get(date_str)

            # C列（業務内容・祝日名など）の記入
            if holiday_name:
                ws[f"C{row}"] = holiday_name
            elif current_date.weekday() < 5:
                ws[f"C{row}"] = task
            else:
                ws[f"C{row}"] = ""

            # 勤務時間の処理（勤務データが存在する場合）
            if not day_data.empty:
                start_time = day_data["Work start"].min()
                end_time = day_data["Work end"].max()
                ws[f"H{row}"] = start_time.time()
                ws[f"I{row}"] = end_time.time()

                valid_breaks = day_data.dropna(subset=["Break start", "Break end"])
                total_break_duration = (valid_breaks["Break end"] - valid_breaks["Break start"]).sum()

                # 追加の休憩時間を加算
                for additional_break in additional_breaks:
                    if additional_break["date"] == date_str and "hours" in additional_break:
                        additional_break_minutes = int(additional_break["hours"] * 60)
                        total_break_duration += pd.Timedelta(minutes=additional_break_minutes)

                total_work_duration = (end_time - start_time) - total_break_duration

                if time_mode == "ratio":
                    ws[f"K{row}"] = ratio_day_fraction
                    # G列：実際の就業時間 - K列の割合固定時間の差分
                    actual_work_hours_decimal = total_work_duration.total_seconds() / 86400
                    ws[f"G{row}"] = actual_work_hours_decimal - ratio_day_fraction
                elif time_mode == "fixed":
                    ws[f"K{row}"] = fixed_hours_fraction
                    # G列：実際の就業時間 - 固定時間の差分
                    actual_work_hours_decimal = total_work_duration.total_seconds() / 86400
                    ws[f"G{row}"] = actual_work_hours_decimal - fixed_hours_fraction
                else:
                    ws[f"K{row}"] = total_work_duration.total_seconds() / 86400  # 実働時間
                    ws[f"G{row}"] = None

                break_minutes = int(total_break_duration.total_seconds() // 60)
                ws[f"J{row}"] = break_minutes / 1440

                for col in ["H", "I", "J", "K", "G"]:
                    if ws[f"{col}{row}"].value is not None:
                        ws[f"{col}{row}"].number_format = "h:mm"

            elif current_date.weekday() < 5 and not holiday_name:
                ws[f"C{row}"] = "休暇"

        # セル検索用の関数を定義
        def find_cell_by_value(ws, value, column=None):
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == value and (column is None or cell.column == column):
                        return cell
            return None

        # 月の日数に応じて不要な日付行を削除
        if days_in_month < 31:
            start_delete_row = 12 + days_in_month + 1
            end_delete_row = 12 + 31
            
            for row_to_delete in range(end_delete_row, start_delete_row - 1, -1):
                ws.delete_rows(row_to_delete)
            
            try:
                actual_day_cell = find_cell_by_value(ws, "実働日")
                if actual_day_cell:
                    merge_range = f"A{actual_day_cell.row}:B{actual_day_cell.row}"
                    ws.merge_cells(merge_range)
            except:
                pass

        # サマリ部分の計算式設定
        end_row = 12 + days_in_month

        work_time_cell = find_cell_by_value(ws, "就業時間", column=5)
        if work_time_cell:
            target = ws.cell(row=work_time_cell.row, column=6)
            target.value = f"=SUM(K13:K{end_row})/TIME(1,,)"
            target.number_format = "0.0"
            
            overtime_target = ws.cell(row=work_time_cell.row, column=9)
            overtime_target.value = f"=F{work_time_cell.row}-C{work_time_cell.row}*8"
            overtime_target.number_format = "0.00"

        actual_day_cell = find_cell_by_value(ws, "実働日")
        if actual_day_cell:
            target = ws.cell(row=actual_day_cell.row, column=3)
            target.value = f"=COUNTA(H13:H{end_row})"

        # ファイル出力処理
        safe_eid = eid.replace(" ", "_").replace("　", "_")
        output_filename = f"タイムシート({year:04d}_{month:02d})_{safe_eid}.xlsx"
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        response = func.HttpResponse(
            body=output_stream.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            status_code=200
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{output_filename}"'
        return add_cors_headers(response)
        
    except Exception as e:
        logging.error(f"Error in generate_timesheet: {str(e)}")
        logging.exception("Full traceback:")
        response = func.HttpResponse(f"エラーが発生しました: {str(e)}", status_code=500)
        return add_cors_headers(response)

@app.route(route="holidays", methods=["GET"])
def get_holidays(req: func.HttpRequest) -> func.HttpResponse:
    """
    祝日データをJSON形式で返すAPI
    """
    year = req.params.get('year')
    logging.info(f"Holidays API accessed with year parameter: {year}")
    
    try:
        holidays_df = pd.read_csv(os.path.join(BASE_DIR, 'holidays.csv'))
        
        if year:
            year = int(year)
            holidays_df['date'] = pd.to_datetime(holidays_df['date'])
            holidays_df = holidays_df[holidays_df['date'].dt.year == year]
        
        holidays_list = []
        for _, row in holidays_df.iterrows():
            holidays_list.append({
                'date': row['date'],
                'name': row['name']
            })
        
        logging.info(f"Returning {len(holidays_list)} holidays for year {year}")
        response = func.HttpResponse(
            body='{"holidays":' + str(holidays_list).replace("'", '"') + '}',
            mimetype="application/json",
            status_code=200
        )
        return add_cors_headers(response)
        
    except Exception as e:
        logging.error(f"Error in holidays API: {str(e)}")
        response = func.HttpResponse(
            body='{"error":"' + str(e) + '"}',
            mimetype="application/json",
            status_code=500
        )
        return add_cors_headers(response)

@app.route(route="holidays-ui", methods=["GET"])
def holidays_ui(req: func.HttpRequest) -> func.HttpResponse:
    """
    祝日管理画面を表示するエンドポイント
    """
    try:
        with open(os.path.join(BASE_DIR, "templates", "holidays.html"), "r", encoding="utf-8") as f:
            html_content = f.read()
        response = func.HttpResponse(
            body=html_content,
            mimetype="text/html",
            status_code=200
        )
        return add_cors_headers(response)
    except Exception as e:
        response = func.HttpResponse(f"Error loading holidays UI: {str(e)}", status_code=500)
        return add_cors_headers(response)

@app.route(route="holidays/download", methods=["GET"])
def download_holidays(req: func.HttpRequest) -> func.HttpResponse:
    """
    祝日データ（holidays.csv）をダウンロードするエンドポイント
    """
    try:
        with open(os.path.join(BASE_DIR, "holidays.csv"), "rb") as f:
            csv_content = f.read()
        response = func.HttpResponse(
            body=csv_content,
            mimetype="text/csv",
            status_code=200
        )
        response.headers['Content-Disposition'] = 'attachment; filename="holidays.csv"'
        return add_cors_headers(response)
    except:
        response = func.HttpResponse("holidays.csv が見つかりません", status_code=404)
        return add_cors_headers(response)

@app.route(route="holidays/upload", methods=["POST", "OPTIONS"])
def upload_holidays(req: func.HttpRequest) -> func.HttpResponse:
    """
    祝日データ（holidays.csv）をアップロードするエンドポイント
    """
    if req.method == "OPTIONS":
        response = func.HttpResponse(status_code=200)
        return add_cors_headers(response)
    
    try:
        file = req.files.get("file")
        if file and file.filename.endswith(".csv"):
            with open(os.path.join(BASE_DIR, "holidays.csv"), "wb") as f:
                f.write(file.stream.read())
            response = func.HttpResponse("アップロード完了", status_code=200)
        else:
            response = func.HttpResponse("CSVファイルのみアップロード可能です", status_code=400)
        return add_cors_headers(response)
    except Exception as e:
        response = func.HttpResponse(f"アップロードエラー: {str(e)}", status_code=500)
        return add_cors_headers(response)
